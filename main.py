# main.py

import sys
import os
import time
import datetime
from collections import Counter
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QLabel,
    QLineEdit, QPushButton, QListWidget, QListWidgetItem, QCalendarWidget,
    QScrollArea, QCheckBox, QToolTip, QDialog, QFormLayout, QTextEdit,
    QDateEdit, QDialogButtonBox, QMenu, QFrame, QMessageBox, QDateTimeEdit,
    QFileDialog, QSizePolicy, QStackedWidget, QComboBox, QTextBrowser
)
from PyQt6.QtGui import (
    QIcon, QFont, QPalette, QColor, QPainter, QCursor, QTextCursor
)
from PyQt6.QtCore import (
    Qt, QSize, pyqtSignal, QDate, QPropertyAnimation, QEasingCurve, QDateTime,
    QParallelAnimationGroup, QAbstractAnimation, QPoint, QTimer
)
from database import DatabaseManager, PRIORITIES, STATUSES, WELCOME_NOTE_TITLE

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import markdown
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False

# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---
CONFIG_FILE = "settings.conf"

def clear_layout(layout):
    if layout:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
            else:
                clear_layout(item.layout())

def load_icon(icon_path):
    if os.path.exists(icon_path):
        return QIcon(icon_path)
    return QIcon()

def colorize_icon(icon: QIcon, color: QColor) -> QIcon:
    if icon.isNull():
        return icon
    pixmap = icon.pixmap(QSize(16, 16))
    painter = QPainter(pixmap)
    painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
    painter.fillRect(pixmap.rect(), color)
    painter.end()
    return QIcon(pixmap)

def apply_stylesheet(app, theme_file):
    try:
        with open(theme_file, "r", encoding="utf-8") as f:
            app.setStyleSheet(f.read())
            return True
    except FileNotFoundError:
        print(f"Внимание: Файл стиля {theme_file} не найден.")
        return False

def save_theme_setting(theme_file):
    with open(CONFIG_FILE, 'w') as f:
        f.write(f"theme={theme_file}")

def load_theme_setting():
    try:
        with open(CONFIG_FILE, 'r') as f:
            line = f.readline()
            if line.startswith("theme="):
                return line.strip().split('=')[1]
    except FileNotFoundError:
        pass
    return "style.qss" # Тема по умолчанию
# ---------------------------------------------


# --- Классы виджетов ---
class MarkdownToolbar(QWidget):
    def __init__(self, text_edit):
        super().__init__()
        self.text_edit = text_edit
        self.setObjectName("MarkdownToolbar")
        
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        buttons_data = [
            ("B", "Жирный", self.on_bold),
            ("I", "Курсив", self.on_italic),
            ("S", "Зачеркнутый", self.on_strikethrough),
            ("H", "Заголовок", self.on_heading),
            ("•", "Список", self.on_list),
            ("</>", "Код", self.on_code),
        ]
        
        for text, tooltip, slot in buttons_data:
            btn = QPushButton(text)
            btn.setToolTip(tooltip)
            btn.setFixedSize(32, 28)
            btn.setObjectName("ToolbarButton")
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.clicked.connect(slot)
            layout.addWidget(btn)
            
        layout.addStretch()
        self.setLayout(layout)

    def wrap_selection(self, prefix, suffix=""):
        cursor = self.text_edit.textCursor()
        if not cursor.hasSelection():
            cursor.insertText(prefix + suffix)
            if suffix:
                cursor.movePosition(QTextCursor.MoveOperation.Left, QTextCursor.MoveMode.MoveAnchor, len(suffix))
            self.text_edit.setTextCursor(cursor)
        else:
            selected_text = cursor.selectedText()
            cursor.insertText(prefix + selected_text + suffix)
        self.text_edit.setFocus()

    def on_bold(self): self.wrap_selection("**", "**")
    def on_italic(self): self.wrap_selection("*", "*")
    def on_strikethrough(self): self.wrap_selection("~~", "~~")
    def on_heading(self): self.wrap_selection("\n## ", "")
    def on_list(self): self.wrap_selection("\n- ", "")
    def on_code(self): self.wrap_selection("\n```python\n", "\n```\n")

class AboutDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("О приложении")
        self.setFixedSize(350, 200)
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        title_label = QLabel("Denkwürfel")
        title_label.setFont(QFont("Inter", 22, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        version_label = QLabel("Версия 1.0")
        version_label.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        help_text = "<p>Менеджер задач и заметок.</p>"
        help_label = QLabel(help_text)
        help_label.setWordWrap(True)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(self.accept)
        layout.addWidget(title_label)
        layout.addWidget(version_label)
        layout.addWidget(separator)
        layout.addWidget(help_label)
        layout.addStretch()
        layout.addWidget(buttons)

class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки")
        self.setMinimumWidth(350)
        
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()

        self.theme_combo = QComboBox()
        if os.path.exists("style.qss"):
            self.theme_combo.addItem("Светлая тема", "style.qss")
        if os.path.exists("dark_style.qss"):
            self.theme_combo.addItem("Темная тема", "dark_style.qss")

        current_theme = load_theme_setting()
        index = self.theme_combo.findData(current_theme)
        if index != -1:
            self.theme_combo.setCurrentIndex(index)

        form_layout.addRow("Тема интерфейса:", self.theme_combo)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.button(QDialogButtonBox.StandardButton.Apply).clicked.connect(self.apply_settings)
        button_box.accepted.connect(self.accept_settings)
        button_box.rejected.connect(self.reject)

        layout.addLayout(form_layout)
        layout.addWidget(button_box)
    
    def apply_settings(self):
        theme_file = self.theme_combo.currentData()
        if theme_file and theme_file != load_theme_setting():
            apply_stylesheet(QApplication.instance(), theme_file)
            save_theme_setting(theme_file)
            if isinstance(self.parent(), MainWindow):
                self.parent().theme_has_changed()
    
    def accept_settings(self):
        self.apply_settings()
        self.accept()

class AddTaskDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить новую задачу")
        self.setMinimumWidth(400)
        self.layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.title_edit = QLineEdit()
        self.details_edit = QTextEdit()
        self.details_edit.setAcceptRichText(False)
        self.tags_edit = QLineEdit()
        self.tags_edit.setPlaceholderText("Например: Работа, Личное, Дом")
        self.due_date_edit = QDateEdit(self)
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDate(QDate.currentDate())
        
        self.priority_combo = QComboBox()
        for i, name in PRIORITIES.items():
            self.priority_combo.addItem(name, i)
        
        self.status_combo = QComboBox()
        self.status_combo.addItems([s for s in STATUSES if s != "Завершено"])

        form_layout.addRow("Название:", self.title_edit)
        form_layout.addRow("Детали:", self.details_edit)
        form_layout.addRow("Теги (через запятую):", self.tags_edit)
        form_layout.addRow("Срок выполнения:", self.due_date_edit)
        form_layout.addRow("Приоритет:", self.priority_combo)
        form_layout.addRow("Статус:", self.status_combo)
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        self.layout.addLayout(form_layout)
        self.layout.addWidget(button_box)
        
    def get_task_data(self):
        return {
            "title": self.title_edit.text().strip(),
            "details": self.details_edit.toPlainText().strip(),
            "tags": self.tags_edit.text().strip(),
            "due_date": self.due_date_edit.date().toPyDate().isoformat(),
            "priority": self.priority_combo.currentData(),
            "status": self.status_combo.currentText()
        }

class EditTaskDialog(QDialog):
    def __init__(self, task_data, reminders, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактировать задачу")
        self.setMinimumWidth(450)
        self.layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        
        self.title_edit = QLineEdit()
        self.details_edit = QTextEdit()
        self.details_edit.setAcceptRichText(False)
        self.tags_edit = QLineEdit()
        self.due_date_edit = QDateEdit(self)
        self.due_date_edit.setCalendarPopup(True)
        
        self.priority_combo = QComboBox()
        for i, name in PRIORITIES.items():
            self.priority_combo.addItem(name, i)
            
        self.status_combo = QComboBox()
        self.status_combo.addItems(STATUSES)

        form_layout.addRow("Название:", self.title_edit)
        form_layout.addRow("Детали:", self.details_edit)
        form_layout.addRow("Теги (через запятую):", self.tags_edit)
        form_layout.addRow("Срок выполнения:", self.due_date_edit)
        form_layout.addRow("Приоритет:", self.priority_combo)
        form_layout.addRow("Статус:", self.status_combo)
        
        self.layout.addLayout(form_layout)
        
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        self.layout.addWidget(separator)
        self.layout.addWidget(QLabel("<b>Напоминания</b>"))
        
        self.reminders_list = QListWidget()
        self.reminders_list.setToolTip("Двойной клик для удаления напоминания.")
        self.reminders_list.itemDoubleClicked.connect(self.remove_selected_reminder)
        self.layout.addWidget(self.reminders_list)
        
        reminder_controls_layout = QHBoxLayout()
        self.reminder_datetime_edit = QDateTimeEdit(self)
        self.reminder_datetime_edit.setCalendarPopup(True)
        self.reminder_datetime_edit.setDateTime(QDateTime.currentDateTime().addSecs(3600))
        self.reminder_datetime_edit.setDisplayFormat("dd.MM.yyyy HH:mm")
        add_reminder_btn = QPushButton("Добавить")
        add_reminder_btn.clicked.connect(self.add_reminder_to_list)
        reminder_controls_layout.addWidget(self.reminder_datetime_edit, 1)
        reminder_controls_layout.addWidget(add_reminder_btn)
        self.layout.addLayout(reminder_controls_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        self.layout.addWidget(button_box)
        
        self.populate_data(task_data, reminders)
        
    def populate_data(self, task_data, reminders):
        self.title_edit.setText(task_data.get('title', ''))
        self.details_edit.setText(task_data.get('details', ''))
        self.tags_edit.setText(task_data.get('tags', ''))
        if due_date_str := task_data.get('due_date'):
            self.due_date_edit.setDate(QDate.fromString(due_date_str, "yyyy-MM-dd"))
            
        self.priority_combo.setCurrentIndex(task_data.get('priority', 0))
        self.status_combo.setCurrentText(task_data.get('status', 'К выполнению'))
        
        for reminder in reminders:
            dt = QDateTime.fromString(reminder['reminder_datetime'], Qt.DateFormat.ISODate)
            item = QListWidgetItem(dt.toString("dd MMMM yy 'в' HH:mm"))
            item.setData(Qt.ItemDataRole.UserRole, reminder['reminder_datetime'])
            self.reminders_list.addItem(item)
            
    def add_reminder_to_list(self):
        dt = self.reminder_datetime_edit.dateTime()
        iso_string = dt.toString(Qt.DateFormat.ISODate)
        for i in range(self.reminders_list.count()):
            if self.reminders_list.item(i).data(Qt.ItemDataRole.UserRole) == iso_string: return
        item = QListWidgetItem(dt.toString("dd MMMM yy 'в' HH:mm"))
        item.setData(Qt.ItemDataRole.UserRole, iso_string)
        self.reminders_list.addItem(item)
        self.reminders_list.sortItems()
        
    def remove_selected_reminder(self, item):
        self.reminders_list.takeItem(self.reminders_list.row(item))
        
    def get_task_data(self):
        return {
            "title": self.title_edit.text().strip(),
            "details": self.details_edit.toPlainText().strip(),
            "tags": self.tags_edit.text().strip(),
            "due_date": self.due_date_edit.date().toPyDate().isoformat(),
            "priority": self.priority_combo.currentData(),
            "status": self.status_combo.currentText()
        }
        
    def get_reminders_data(self):
        return [self.reminders_list.item(i).data(Qt.ItemDataRole.UserRole) for i in range(self.reminders_list.count())]

class ReportDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки отчета по задачам")
        self.setMinimumWidth(400)
        
        self.layout = QVBoxLayout(self)
        form_layout = QFormLayout()

        self.filter_combo = QComboBox()
        self.filter_combo.addItem("Все задачи (активные и завершенные)", "all_statuses")
        self.filter_combo.addItem("Только активные задачи", "active")
        self.filter_combo.addItem("Только завершенные задачи", "completed")
        self.filter_combo.addItem("Задачи с высоким приоритетом", "important")

        self.date_range_check = QCheckBox("Ограничить отчет по дате")
        self.date_range_check.stateChanged.connect(self.toggle_date_fields)
        
        self.start_date_label = QLabel("Начальная дата:")
        self.start_date_edit = QDateEdit(self, calendarPopup=True, date=QDate.currentDate().addDays(-7))
        self.end_date_label = QLabel("Конечная дата:")
        self.end_date_edit = QDateEdit(self, calendarPopup=True, date=QDate.currentDate())
        
        self.sort_combo = QComboBox()
        self.sort_combo.addItem("Сортировать по приоритету", "priority")
        self.sort_combo.addItem("Сортировать по сроку выполнения", "due_date")
        self.sort_combo.addItem("Сортировать по дате создания", "creation_date")
        self.sort_combo.addItem("Сортировать по алфавиту", "alphabetical")

        form_layout.addRow("Включить в отчет:", self.filter_combo)
        form_layout.addRow(self.date_range_check)
        form_layout.addRow(self.start_date_label, self.start_date_edit)
        form_layout.addRow(self.end_date_label, self.end_date_edit)
        form_layout.addRow("Отсортировать по:", self.sort_combo)
        
        self.layout.addLayout(form_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        self.layout.addWidget(button_box)
        
        self.toggle_date_fields()

    def toggle_date_fields(self):
        is_checked = self.date_range_check.isChecked()
        self.start_date_label.setVisible(is_checked)
        self.start_date_edit.setVisible(is_checked)
        self.end_date_label.setVisible(is_checked)
        self.end_date_edit.setVisible(is_checked)
        
    def get_report_settings(self):
        return {
            "filter_by": self.filter_combo.currentData(),
            "use_date_range": self.date_range_check.isChecked(),
            "sort_by": self.sort_combo.currentData(),
            "start_date": self.start_date_edit.date().toPyDate().isoformat(),
            "end_date": self.end_date_edit.date().toPyDate().isoformat()
        }

class TaskWidget(QWidget):
    edit_requested = pyqtSignal(int)
    
    def __init__(self, task_data):
        super().__init__()
        self.task_id = task_data['id']
        self.setObjectName("TaskWidget")
        self.setWindowOpacity(0.0)
        self.setMaximumHeight(0)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 10, 10, 10)
        layout.setSpacing(10)

        self.priority_indicator = QFrame()
        self.priority_indicator.setObjectName("PriorityIndicator")
        self.priority_indicator.setFixedWidth(5)
        self.priority_indicator.setProperty("priority", task_data.get('priority', 0))
        
        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)
        title_label = QLabel(task_data['title'])
        title_label.setObjectName("TaskTitle")
        title_label.setWordWrap(True)

        meta_text = []
        if task_data['tags']:
            meta_text.append(f"🏷️ {task_data['tags']}")
        if task_data['due_date']:
            try:
                meta_text.append(f"🗓️ {datetime.date.fromisoformat(task_data['due_date']).strftime('%d %b')}")
            except (ValueError, TypeError): pass
        
        meta_label = QLabel("  ".join(meta_text))
        meta_label.setObjectName("TaskMeta")
        text_layout.addWidget(title_label)
        if meta_text: text_layout.addWidget(meta_label)
        
        layout.addWidget(self.priority_indicator)
        layout.addLayout(text_layout, 1)
        
        self.status_label = QLabel(task_data['status'])
        self.status_label.setObjectName("StatusLabel")
        if task_data['status'] == 'К выполнению' or task_data['status'] == 'Завершено':
            self.status_label.hide()
        layout.addWidget(self.status_label)

        if task_data['details']:
            self.setToolTip(f"<b>Детали:</b><br>{task_data['details']}")
            
        self.update_visual_state(task_data['status'])

    def update_visual_state(self, status):
        is_completed = (status == 'Завершено')
        self.setProperty("completed", is_completed)
        self.style().unpolish(self)
        self.style().polish(self)
    
    def mouseDoubleClickEvent(self, event):
        self.edit_requested.emit(self.task_id)
        super().mouseDoubleClickEvent(event)
    
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.current_task_filter = 'active'
        self.current_task_filter_value = None
        self.current_sort_by = 'priority'
        self.current_title = "Активные задачи"
        self.active_animations = []
        
        self.setWindowTitle("Denkwürfel")
        self.setGeometry(100, 100, 1280, 800)
        
        self.update_icons()
        
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.init_ui(main_layout)
        
        self.refresh_all_views()
        self.on_main_nav_clicked(self.main_nav_list.item(0))

        self.reminder_timer = QTimer(self)
        self.reminder_timer.timeout.connect(self.check_for_reminders)
        self.reminder_timer.start(30000)

    def update_icons(self):
        text_color = self.palette().color(QPalette.ColorRole.Text)
        raw_icons = {
            "tasks": load_icon("icons/completed.svg"), 
            "notes": load_icon("icons/personal.svg"), 
            "important": load_icon("icons/important.svg"), 
            "all": load_icon("icons/tag.svg"), 
            "completed": load_icon("icons/completed.svg"), 
            "tag": load_icon("icons/tag.svg"),
            "settings": load_icon("icons/settings.svg")
        }
        self.icons = {name: colorize_icon(icon, text_color) for name, icon in raw_icons.items()}

    def init_ui(self, main_layout):
        left_panel = self.create_left_panel()
        center_panel = self.create_center_panel()
        right_panel = self.create_right_panel()
        main_layout.addWidget(left_panel)
        main_layout.addWidget(center_panel, 1)
        main_layout.addWidget(right_panel)

    def create_left_panel(self):
        left_panel = QWidget()
        left_panel.setObjectName("LeftPanel")
        left_panel.setFixedWidth(250)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(10, 10, 10, 10)
        left_layout.setSpacing(5) 
        title_label = QLabel("Denkwürfel")
        title_label.setObjectName("AppTitle")
        
        self.main_nav_list = QListWidget()
        self.main_nav_list.setObjectName("NavList")
        self.main_nav_list.addItem(QListWidgetItem(self.icons.get("tasks"), "Задачи"))
        self.main_nav_list.addItem(QListWidgetItem(self.icons.get("notes"), "Заметки"))
        self.main_nav_list.itemClicked.connect(self.on_main_nav_clicked)
        
        self.main_nav_list.setSizeAdjustPolicy(QListWidget.SizeAdjustPolicy.AdjustToContents)
        self.main_nav_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        
        self.left_panel_stack = QStackedWidget()
        
        tasks_nav_widget = QWidget()
        tasks_nav_widget.setObjectName("TasksNavWidget")
        tasks_nav_layout = QVBoxLayout(tasks_nav_widget)
        tasks_nav_layout.setContentsMargins(0,0,0,0)
        tasks_nav_layout.setSpacing(5)

        self.search_bar = QLineEdit(placeholderText="🔍 Поиск по задачам")
        self.search_bar.setObjectName("SearchBar")
        self.search_bar.textChanged.connect(self.on_search_text_changed)
        
        filters_label = QLabel("Фильтры")
        filters_label.setObjectName("NavHeaderLabel")
        self.task_filters_list = QListWidget()
        self.task_filters_list.setObjectName("NavList")
        self.task_filters_list.itemClicked.connect(self.on_task_filter_clicked)
        
        tags_label = QLabel("Теги")
        tags_label.setObjectName("NavHeaderLabel")
        self.tags_list = QListWidget()
        self.tags_list.setObjectName("NavList")
        self.tags_list.itemClicked.connect(self.on_tag_item_clicked)
        
        self.report_button = QPushButton("Выгрузить отчет")
        self.report_button.setObjectName("ReportButton")
        self.report_button.clicked.connect(self.show_report_dialog)
        
        tasks_nav_layout.addWidget(self.search_bar)
        tasks_nav_layout.addWidget(filters_label)
        tasks_nav_layout.addWidget(self.task_filters_list)
        tasks_nav_layout.addWidget(tags_label)
        tasks_nav_layout.addWidget(self.tags_list)
        tasks_nav_layout.addStretch(1)
        tasks_nav_layout.addWidget(self.report_button)

        notes_nav_widget = QWidget()
        notes_nav_layout = QVBoxLayout(notes_nav_widget)
        notes_nav_layout.setContentsMargins(0,0,0,0)
        self.notes_list_widget = QListWidget()
        self.notes_list_widget.setObjectName("NavList")
        self.notes_list_widget.itemDoubleClicked.connect(self.open_note_in_editor)
        notes_nav_layout.addWidget(QLabel("Ваши заметки"))
        notes_nav_layout.addWidget(self.notes_list_widget, 1)
        
        self.left_panel_stack.addWidget(tasks_nav_widget)
        self.left_panel_stack.addWidget(notes_nav_widget)
        
        left_layout.addWidget(title_label)
        left_layout.addWidget(self.main_nav_list)
        left_layout.addWidget(self.left_panel_stack, 1)
        return left_panel

    def create_center_panel(self):
        center_panel = QWidget()
        center_panel.setObjectName("CenterPanel")
        self.center_layout = QVBoxLayout(center_panel)
        self.center_layout.setContentsMargins(20, 20, 20, 20)
        
        header_layout = QHBoxLayout()
        self.center_title_label = QLabel(self.current_title)
        self.center_title_label.setObjectName("CenterTitle")
        
        self.sort_combo = QComboBox()
        self.sort_combo.setObjectName("SortComboBox")
        self.sort_combo.addItem("Сортировать по приоритету", "priority")
        self.sort_combo.addItem("Сортировать по сроку", "due_date")
        self.sort_combo.addItem("Сортировать по дате создания", "creation_date")
        self.sort_combo.addItem("Сортировать по алфавиту", "alphabetical")
        self.sort_combo.currentIndexChanged.connect(self.on_sort_changed)
        
        self.new_item_button = QPushButton("+ Новая задача")
        self.new_item_button.setObjectName("NewTaskButton")
        self.new_item_button.clicked.connect(self.on_new_item_button_clicked)
        
        header_layout.addWidget(self.center_title_label)
        header_layout.addStretch()
        header_layout.addWidget(self.sort_combo)
        header_layout.addWidget(self.new_item_button)
        
        self.center_stack = QStackedWidget()
        self.view_widget = QWidget()
        self.view_layout = QVBoxLayout(self.view_widget)
        self.view_layout.setContentsMargins(0,0,0,0)
        self.editor_widget = self.create_note_editor_view()
        self.center_stack.addWidget(self.view_widget)
        self.center_stack.addWidget(self.editor_widget)
        
        self.center_layout.addLayout(header_layout)
        self.center_layout.addWidget(self.center_stack, 1)
        return center_panel
        
    def create_note_editor_view(self):
        editor_container = QWidget()
        layout = QVBoxLayout(editor_container)
        layout.setContentsMargins(0, 10, 0, 0)
        layout.setSpacing(5)

        self.note_title_edit = QLineEdit()
        self.note_title_edit.setObjectName("NoteTitleEdit")
        
        self.note_content_edit = QTextEdit()
        self.note_content_edit.setObjectName("NoteContentEdit")
        
        self.markdown_toolbar = MarkdownToolbar(self.note_content_edit)
        
        buttons_layout = QHBoxLayout()
        
        self.delete_note_button = QPushButton("Удалить")
        self.delete_note_button.setObjectName("DeleteNoteButton")
        
        self.save_note_button = QPushButton("Сохранить")
        self.save_note_button.setObjectName("SaveNoteButton")
        
        self.close_editor_button = QPushButton("Закрыть")
        self.close_editor_button.setObjectName("CloseEditorButton")

        self.delete_note_button.clicked.connect(self.delete_current_note)
        self.save_note_button.clicked.connect(self.save_current_note)
        self.close_editor_button.clicked.connect(self.close_note_editor)
        
        buttons_layout.addWidget(self.delete_note_button)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.close_editor_button)
        buttons_layout.addWidget(self.save_note_button)
        
        layout.addWidget(self.note_title_edit)
        layout.addWidget(self.markdown_toolbar)
        layout.addWidget(self.note_content_edit, 1)
        layout.addLayout(buttons_layout)
        return editor_container

    def create_right_panel(self):
        right_panel = QWidget()
        right_panel.setObjectName("RightPanel")
        right_panel.setFixedWidth(300)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(10, 15, 10, 15)
        
        top_layout = QHBoxLayout()
        top_layout.setContentsMargins(0,0,0,0)
        
        self.settings_button = QPushButton("Настройки")
        self.settings_button.setIcon(self.icons.get("settings"))
        self.settings_button.setObjectName("SettingsButton")
        self.settings_button.clicked.connect(self.show_settings_menu)
        
        top_layout.addStretch()
        top_layout.addWidget(self.settings_button)
        
        right_layout.addLayout(top_layout)
        
        self.right_stack = QStackedWidget()
        
        tasks_right_panel = QWidget()
        tasks_right_layout = QVBoxLayout(tasks_right_panel)
        tasks_right_layout.setContentsMargins(0, 10, 0, 0)
        self.calendar = QCalendarWidget(verticalHeaderFormat=QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader, gridVisible=True)
        self.calendar.setObjectName("CalendarWidget")
        self.calendar.selectionChanged.connect(self.on_date_selected)
        self.completed_label = QLabel("Завершенные задачи")
        self.completed_label.setFont(QFont("Inter", 12, QFont.Weight.Bold))
        self.completed_list_widget = QListWidget()
        self.completed_list_widget.setObjectName("CompletedList")
        tasks_right_layout.addWidget(self.calendar)
        tasks_right_layout.addSpacing(10)
        tasks_right_layout.addWidget(self.completed_label)
        tasks_right_layout.addWidget(self.completed_list_widget)
        
        notes_right_panel = QWidget()
        notes_right_layout = QVBoxLayout(notes_right_panel)
        notes_right_layout.setContentsMargins(0, 10, 0, 0)
        preview_label = QLabel("Предпросмотр")
        preview_label.setFont(QFont("Inter", 12, QFont.Weight.Bold))
        self.markdown_preview = QTextBrowser()
        self.markdown_preview.setObjectName("MarkdownPreview")
        self.markdown_preview.setOpenExternalLinks(True)
        notes_right_layout.addWidget(preview_label)
        notes_right_layout.addWidget(self.markdown_preview)

        self.right_stack.addWidget(tasks_right_panel)
        self.right_stack.addWidget(notes_right_panel)
        
        right_layout.addWidget(self.right_stack)
        
        return right_panel

    def switch_to_tasks_view(self):
        clear_layout(self.view_layout)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setObjectName("ScrollArea")
        tasks_container = QWidget()
        self.tasks_layout = QVBoxLayout(tasks_container)
        self.tasks_layout.setSpacing(0)
        self.tasks_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        scroll_area.setWidget(tasks_container)
        self.view_layout.addWidget(scroll_area)
        self.center_stack.setCurrentWidget(self.view_widget)
        self.sort_combo.show()
        self.refresh_task_list(animated=True)

    def switch_to_notes_view(self):
        clear_layout(self.view_layout)
        welcome_label = QLabel("Выберите заметку слева или создайте новую.\n\nДвойной клик по заметке откроет редактор.")
        welcome_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        welcome_label.setObjectName("WelcomeLabel")
        self.view_layout.addWidget(welcome_label)
        self.center_stack.setCurrentWidget(self.view_widget)
        self.sort_combo.hide()
    
    def animate_show_item(self, widget, duration):
        group = QParallelAnimationGroup(self)
        opacity_anim = QPropertyAnimation(widget, b"windowOpacity")
        opacity_anim.setDuration(duration)
        opacity_anim.setStartValue(0.0)
        opacity_anim.setEndValue(1.0)
        size_anim = QPropertyAnimation(widget, b"maximumHeight")
        size_anim.setDuration(duration)
        size_anim.setStartValue(0)
        size_anim.setEndValue(widget.sizeHint().height())
        size_anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        group.addAnimation(opacity_anim)
        group.addAnimation(size_anim)
        group.finished.connect(lambda: self.active_animations.remove(group) if group in self.active_animations else None)
        self.active_animations.append(group)
        group.start(QAbstractAnimation.DeletionPolicy.DeleteWhenStopped)

    def on_sort_changed(self, index):
        self.current_sort_by = self.sort_combo.itemData(index)
        self.refresh_task_list(animated=True)

    def refresh_all_views(self):
        self.refresh_task_filters_list()
        self.refresh_tags_list()
        if hasattr(self, 'tasks_layout'):
            self.refresh_task_list()
        self.refresh_completed_list()
        self.refresh_notes_list()

    def refresh_main_views(self, animated=False):
        self.refresh_tags_list()
        if hasattr(self, 'tasks_layout'):
            self.refresh_task_list(animated=animated)
        self.refresh_completed_list()

    def refresh_task_list(self, animated=False, tasks_list=None):
        if not hasattr(self, 'tasks_layout'): return
        clear_layout(self.tasks_layout)
        tasks = tasks_list if tasks_list is not None else self.db.get_tasks(
            filter_by=self.current_task_filter, 
            value=self.current_task_filter_value,
            sort_by=self.current_sort_by
        )
        for i, task_data in enumerate(tasks):
            task_widget = TaskWidget(task_data)
            task_widget.edit_requested.connect(self.show_edit_task_dialog)
            self.tasks_layout.addWidget(task_widget)
            if animated: self.animate_show_item(task_widget, 250 + i * 25)

    def refresh_task_filters_list(self):
        self.task_filters_list.clear()
        self.task_filters_list.addItem(QListWidgetItem(self.icons.get("all"), "Активные задачи"))
        self.task_filters_list.addItem(QListWidgetItem(self.icons.get("important"), "Высокий приоритет"))
        self.task_filters_list.addItem(QListWidgetItem(self.icons.get("completed"), "Завершенные"))
    
    def refresh_tags_list(self):
        self.tags_list.clear()
        for tag, count in sorted(self.db.get_tags_with_counts().items()):
            item = QListWidgetItem(self.tags_list)
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(5, 3, 8, 3)
            row_layout.setSpacing(6)
            icon_label = QLabel()
            icon_label.setPixmap(self.icons.get("tag").pixmap(QSize(16, 16)))
            count_label = QLabel(str(count))
            count_label.setObjectName("TagCount")
            count_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            row_layout.addWidget(icon_label)
            row_layout.addWidget(QLabel(tag), 1)
            row_layout.addWidget(count_label)
            item.setData(Qt.ItemDataRole.UserRole, tag)
            self.tags_list.setItemWidget(item, row_widget)
            
    def refresh_completed_list(self):
        self.completed_list_widget.clear()
        for task in self.db.get_tasks(filter_by='completed')[:5]:
            item = QListWidgetItem(f"✔ {task['title']}")
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
            self.completed_list_widget.addItem(item)
    
    def refresh_notes_list(self):
        self.db._ensure_welcome_note_exists()
        current_selection = self.notes_list_widget.currentItem()
        current_id = current_selection.data(Qt.ItemDataRole.UserRole) if current_selection else None
        self.notes_list_widget.clear()
        item_to_reselect = None
        for note in self.db.get_all_notes():
            item = QListWidgetItem(note['title'])
            item.setData(Qt.ItemDataRole.UserRole, note['id'])
            self.notes_list_widget.addItem(item)
            if note['id'] == current_id: item_to_reselect = item
        if item_to_reselect: self.notes_list_widget.setCurrentItem(item_to_reselect)

    def on_main_nav_clicked(self, item):
        if item.text() == "Задачи":
            self.left_panel_stack.setCurrentIndex(0)
            self.new_item_button.setText("+ Новая задача")
            self.center_title_label.setText("Активные задачи")
            self.current_task_filter = 'active'
            self.switch_to_tasks_view()
            self.right_stack.setCurrentIndex(0)
        elif item.text() == "Заметки":
            self.left_panel_stack.setCurrentIndex(1)
            self.new_item_button.setText("+ Новая заметка")
            self.center_title_label.setText("Заметки")
            self.switch_to_notes_view()
            self.right_stack.setCurrentIndex(1)

    def on_task_filter_clicked(self, item):
        self.search_bar.clear()
        filter_text = item.text()
        self.current_title = filter_text
        self.current_task_filter_value = None
        if filter_text == "Активные задачи": self.current_task_filter = 'active'
        elif filter_text == "Высокий приоритет": self.current_task_filter = 'important'
        elif filter_text == "Завершенные": self.current_task_filter = 'completed'
        self.center_title_label.setText(self.current_title)
        self.refresh_task_list(animated=True)
    
    def on_tag_item_clicked(self, item):
        self.search_bar.clear()
        if tag_name := item.data(Qt.ItemDataRole.UserRole):
            self.current_task_filter = 'tag'
            self.current_task_filter_value = tag_name
            self.current_title = f"Тег: {tag_name}"
            self.center_title_label.setText(self.current_title)
            self.refresh_task_list(animated=True)

    def on_date_selected(self):
        self.main_nav_list.setCurrentRow(0)
        self.on_main_nav_clicked(self.main_nav_list.item(0))
        self.search_bar.clear()
        selected_date_q = self.calendar.selectedDate()
        self.current_task_filter = 'date'
        self.current_task_filter_value = selected_date_q.toString("yyyy-MM-dd")
        self.current_title = f"Задачи на {selected_date_q.toString('d MMMM yyyy г.')}"
        self.center_title_label.setText(self.current_title)
        self.refresh_task_list(animated=True)

    def on_search_text_changed(self, text):
        query = text.strip()
        if query:
            self.center_title_label.setText(f'Результаты поиска: "{query}"')
            self.refresh_task_list(animated=True, tasks_list=self.db.search_tasks(query))
        else:
            self.center_title_label.setText(self.current_title)
            self.refresh_task_list(animated=True)

    def on_new_item_button_clicked(self):
        if self.main_nav_list.currentItem().text() == "Задачи":
            menu = QMenu(self)
            menu.addAction("Добавить задачу", lambda: self.show_add_task_dialog())
            menu.addAction("Добавить с высоким приоритетом", lambda: self.show_add_task_dialog(high_priority=True))
            menu.exec(self.new_item_button.mapToGlobal(QPoint(0, self.new_item_button.height())))
        else:
            self.open_note_in_editor(None)
    
    def show_settings_menu(self):
        menu = QMenu(self)
        menu.addAction("О приложении", self.show_about_dialog)
        menu.addAction("Настройки интерфейса", self.show_settings_dialog)
        menu.exec(self.settings_button.mapToGlobal(QPoint(0, self.settings_button.height())))

    def show_about_dialog(self):
        dialog = AboutDialog(self)
        dialog.exec()
        
    def show_settings_dialog(self):
        dialog = SettingsDialog(self)
        dialog.exec()
    
    def theme_has_changed(self):
        self.update_icons()
        self.settings_button.setIcon(self.icons.get("settings"))
        self.refresh_all_views()
        
    def show_add_task_dialog(self, high_priority=False):
        dialog = AddTaskDialog(self)
        if high_priority:
            dialog.priority_combo.setCurrentIndex(3)
        if dialog.exec():
            task_data = dialog.get_task_data()
            if task_data['title']:
                self.db.add_task(**task_data)
                self.refresh_main_views(animated=True)
                
    def show_edit_task_dialog(self, task_id):
        task_data = self.db.get_task_by_id(task_id)
        if not task_data: return
        reminders = self.db.get_reminders_for_task(task_id)
        dialog = EditTaskDialog(task_data, reminders, self)
        if dialog.exec():
            new_data = dialog.get_task_data()
            if new_data['title']:
                self.db.update_task(task_id, new_data)
                self.db.replace_all_reminders_for_task(task_id, dialog.get_reminders_data())
                self.refresh_main_views(animated=True)
    
    def open_note_in_editor(self, item):
        self.current_note_id = item.data(Qt.ItemDataRole.UserRole) if item else None
        
        self.note_title_edit.setReadOnly(False)
        self.note_content_edit.setReadOnly(False)
        self.save_note_button.show()
        self.markdown_toolbar.show()

        if self.current_note_id:
            note_data = self.db.get_note_by_id(self.current_note_id)
            if note_data:
                self.note_title_edit.setText(note_data.get('title', ''))
                self.note_content_edit.setPlainText(note_data.get('content', ''))
                
                is_welcome_note = note_data.get('title') == WELCOME_NOTE_TITLE
                self.note_title_edit.setReadOnly(is_welcome_note)
                self.note_content_edit.setReadOnly(is_welcome_note)
                self.save_note_button.setVisible(not is_welcome_note)
                self.markdown_toolbar.setVisible(not is_welcome_note)
                
                self.delete_note_button.setDisabled(is_welcome_note)
                self.delete_note_button.setToolTip("Эту заметку нельзя удалить" if is_welcome_note else "")
                self.delete_note_button.setVisible(True)
                
                self.note_content_edit.textChanged.connect(self.update_markdown_preview)
                self.update_markdown_preview()
            else:
                QMessageBox.warning(self, "Ошибка", "Заметка не найдена. Возможно, она была удалена.")
                self.close_note_editor()
                self.refresh_notes_list()
        else:
            self.note_title_edit.clear()
            self.note_content_edit.clear()
            self.note_title_edit.setPlaceholderText("Новая заметка...")
            self.delete_note_button.hide()
            self.note_content_edit.textChanged.connect(self.update_markdown_preview)
            self.update_markdown_preview()
        
        self.center_stack.setCurrentWidget(self.editor_widget)
        self.right_stack.setCurrentIndex(1)

    def save_current_note(self):
        title = self.note_title_edit.text().strip()
        if not title:
            QMessageBox.warning(self, "Ошибка", "Заголовок заметки не может быть пустым.")
            return
        
        content = self.note_content_edit.toPlainText()
        
        if self.current_note_id:
            self.db.update_note(self.current_note_id, title, content)
        else:
            self.current_note_id = self.db.add_note(title, content)
        
        self.refresh_notes_list()

    def close_note_editor(self):
        try:
            self.note_content_edit.textChanged.disconnect(self.update_markdown_preview)
        except TypeError:
            pass
        self.markdown_preview.clear()
        self.center_stack.setCurrentWidget(self.view_widget)
        self.current_note_id = None
        self.right_stack.setCurrentIndex(0)

    def delete_current_note(self):
        if not self.current_note_id: return
        reply = QMessageBox.question(self, "Подтверждение", "Вы уверены, что хотите удалить эту заметку?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_note(self.current_note_id)
            self.refresh_notes_list()
            self.close_note_editor()

    def check_for_reminders(self):
        now_iso = datetime.datetime.now().isoformat()
        due_reminders = self.db.get_due_reminders(now_iso)
        for reminder in due_reminders:
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Icon.Information)
            msg_box.setWindowTitle("Напоминание о задаче")
            dt = QDateTime.fromString(reminder['reminder_datetime'], Qt.DateFormat.ISODate)
            msg_box.setText(f"<b>{reminder['title']}</b>")
            msg_box.setInformativeText(f"Время выполнить задачу! (Напоминание на {dt.toString('HH:mm')})")
            msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg_box.exec()
            self.db.delete_reminder(reminder['reminder_id'])

    def _generate_report_summary(self, tasks):
        total_tasks = len(tasks)
        if total_tasks == 0:
            return {"stats": {}}

        completed_tasks = sum(1 for task in tasks if task['status'] == 'Завершено')
        completion_percentage = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        priority_counts = Counter(task['priority'] for task in tasks)
        tag_counts = Counter(tag.strip() for task in tasks if task.get('tags') for tag in task['tags'].split(',') if tag.strip())

        stats = {
            "total": total_tasks,
            "completed": completed_tasks,
            "percentage": completion_percentage,
            "priorities": priority_counts,
            "tags": tag_counts
        }
        
        return {"stats": stats}

    def show_report_dialog(self):
        dialog = ReportDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            settings = dialog.get_report_settings()
            
            start_date = settings["start_date"] if settings["use_date_range"] else None
            end_date = settings["end_date"] if settings["use_date_range"] else None

            report_tasks = self.db.get_tasks(
                filter_by=settings["filter_by"],
                sort_by=settings["sort_by"],
                start_date=start_date,
                end_date=end_date
            )
            
            if not report_tasks:
                QMessageBox.information(self, "Нет данных", "Задачи для отчета не найдены по выбранным критериям.")
                return

            summary = self._generate_report_summary(report_tasks)
            
            filter_name = dialog.filter_combo.currentText()
            default_filename = f"Отчет - {filter_name}"
            if settings["use_date_range"]:
                default_filename += f" ({settings['start_date']} - {settings['end_date']})"

            filters = "Excel Files (*.xlsx)"
            filePath, _ = QFileDialog.getSaveFileName(self, "Сохранить отчет", default_filename, filters)
            
            if filePath:
                self.save_report_as_excel(report_tasks, summary["stats"], filePath)
            
    def save_report_as_excel(self, tasks, summary_stats, file_path):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Ошибка", "Для экспорта в Excel необходимо установить библиотеку openpyxl.\nВыполните: pip install openpyxl")
            return
            
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Отчет по задачам"
        
        bold_font = Font(bold=True)
        
        start_row = 1
        sheet.cell(row=start_row, column=1, value="Сводная статистика").font = bold_font
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        start_row += 1
        
        sheet.cell(row=start_row, column=1, value="Всего задач:")
        sheet.cell(row=start_row, column=2, value=summary_stats.get('total'))
        start_row += 1
        
        sheet.cell(row=start_row, column=1, value="Завершено:")
        sheet.cell(row=start_row, column=2, value=f"{summary_stats.get('completed')} ({summary_stats.get('percentage', 0):.1f}%)")
        start_row += 1
        
        sheet.cell(row=start_row, column=1, value="По приоритетам:").font = bold_font
        start_row +=1
        for prio_id, count in sorted(summary_stats.get('priorities', {}).items(), key=lambda item: item[0], reverse=True):
            sheet.cell(row=start_row, column=2, value=f"{PRIORITIES.get(prio_id, 'Н/Д')}: {count}")
            start_row += 1
            
        sheet.cell(row=start_row, column=1, value="По тегам:").font = bold_font
        start_row += 1
        tag_counts = summary_stats.get('tags')
        if tag_counts:
            for tag, count in tag_counts.most_common():
                sheet.cell(row=start_row, column=2, value=f"{tag}: {count}")
                start_row += 1
        else:
             sheet.cell(row=start_row, column=2, value="Нет")
             start_row += 1

        start_row += 1 
        headers = ["Задача", "Статус", "Приоритет", "Срок выполнения", "Детали", "Теги"]
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col_num, value=header_title)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for task in tasks:
            start_row += 1
            priority_text = PRIORITIES.get(task.get('priority', 0), "Нет")
            due_date = QDate.fromString(task['due_date'], 'yyyy-MM-dd').toString('dd.MM.yyyy') if task.get('due_date') else ""
            sheet.cell(row=start_row, column=1, value=task.get('title', ''))
            sheet.cell(row=start_row, column=2, value=task.get('status', ''))
            sheet.cell(row=start_row, column=3, value=priority_text)
            sheet.cell(row=start_row, column=4, value=due_date)
            sheet.cell(row=start_row, column=5, value=task.get('details', ''))
            sheet.cell(row=start_row, column=6, value=task.get('tags', ''))
            
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in sheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 70)

        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен в файл:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл отчета.\nОшибка: {e}")

    def update_markdown_preview(self):
        if not MARKDOWN_AVAILABLE:
            self.markdown_preview.setText("<b>Ошибка:</b> Библиотека <code>markdown</code> не найдена.<br>Установите ее командой: <code>pip install markdown</code>")
            return
        
        markdown_text = self.note_content_edit.toPlainText()
        html = markdown.markdown(markdown_text, extensions=['fenced_code', 'tables', 'sane_lists'])
        self.markdown_preview.setHtml(html)

    def closeEvent(self, event):
        self.db.close()
        super().closeEvent(event)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    QToolTip.setFont(QFont("Inter", 10))
    if os.path.exists("icons/icons.png"): app.setWindowIcon(QIcon("icons/icons.png"))
    
    saved_theme = load_theme_setting()
    apply_stylesheet(app, saved_theme)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())